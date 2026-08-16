"""Lectura y validación de archivos Excel para la inserción masiva."""

from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path
from typing import Any, Mapping, Sequence

from database import ConfigurationError, ValidationError, convert_value

try:
    from openpyxl import load_workbook
except ModuleNotFoundError:  # Mensaje más claro cuando falta la dependencia.
    load_workbook = None  # type: ignore[assignment]


@dataclass
class ExcelImportData:
    """Archivo ya leído y validado, listo para enviarse al repositorio."""

    file_path: str
    sheet_name: str
    columns: list[str]
    preview_rows: list[dict[str, Any]]
    prepared_rows: list[list[Any]]
    source_row_numbers: list[int]

    @property
    def row_count(self) -> int:
        return len(self.prepared_rows)


class ExcelImportService:
    """Convierte filas de Excel usando la misma configuración del formulario."""

    def __init__(
        self,
        fields: Sequence[Mapping[str, Any]],
        config: Mapping[str, Any],
    ) -> None:
        self.fields = list(fields)
        self.config = dict(config)
        self.field_names = [str(field["name"]) for field in self.fields]
        self.fields_by_normalized = {
            self._normalize(name): field
            for name, field in zip(self.field_names, self.fields)
        }
        self._validate_configuration()

    @staticmethod
    def _normalize(value: Any) -> str:
        return str(value).strip().casefold()

    @staticmethod
    def _is_blank(value: Any) -> bool:
        return value is None or (isinstance(value, str) and not value.strip())

    def _validate_configuration(self) -> None:
        if not self.fields:
            raise ConfigurationError("No hay campos configurados para importar.")
        allowed = self.config.get("allowed_extensions", [".xlsx", ".xlsm"])
        if not isinstance(allowed, (list, tuple)) or not allowed:
            raise ConfigurationError("allowed_extensions debe ser una lista no vacía.")
        for extension in allowed:
            if not str(extension).startswith("."):
                raise ConfigurationError(
                    f"Extensión no válida en allowed_extensions: {extension!r}."
                )
        if int(self.config.get("header_row", 1)) < 1:
            raise ConfigurationError("header_row debe ser mayor o igual que 1.")
        if int(self.config.get("preview_rows", 20)) < 1:
            raise ConfigurationError("preview_rows debe ser mayor o igual que 1.")
        if int(self.config.get("max_rows", 100000)) < 1:
            raise ConfigurationError("max_rows debe ser mayor o igual que 1.")
        batch_size = int(self.config.get("batch_size", 500))
        if not 1 <= batch_size <= 5000:
            raise ConfigurationError("batch_size debe estar entre 1 y 5000.")

        mapping = self.config.get("column_mapping", {})
        if not isinstance(mapping, Mapping):
            raise ConfigurationError("column_mapping debe ser un diccionario.")
        valid_targets = set(self.fields_by_normalized)
        invalid_targets = [
            str(target)
            for target in mapping.values()
            if self._normalize(target) not in valid_targets
        ]
        if invalid_targets:
            raise ConfigurationError(
                "column_mapping contiene columnas SQL no configuradas: "
                + ", ".join(invalid_targets)
            )

    def _choose_sheet(self, workbook: Any) -> Any:
        configured = self.config.get("sheet_name")
        if configured in (None, ""):
            return workbook.active
        sheet_name = str(configured)
        if sheet_name not in workbook.sheetnames:
            raise ValidationError(
                f"La hoja {sheet_name!r} no existe. Disponibles: "
                + ", ".join(workbook.sheetnames)
            )
        return workbook[sheet_name]

    def _resolve_headers(
        self, header_values: Sequence[Any]
    ) -> tuple[dict[int, str], list[str]]:
        expected = {
            self._normalize(name): name
            for name in self.field_names
        }
        configured_mapping = {
            self._normalize(source): expected[self._normalize(target)]
            for source, target in self.config.get("column_mapping", {}).items()
        }
        indexes: dict[int, str] = {}
        extra_headers: list[str] = []
        already_mapped: set[str] = set()

        for index, header in enumerate(header_values):
            if self._is_blank(header):
                continue
            source_name = str(header).strip()
            normalized_source = self._normalize(source_name)
            target_name = configured_mapping.get(normalized_source)
            if target_name is None:
                target_name = expected.get(normalized_source)
            if target_name is None:
                extra_headers.append(source_name)
                continue

            normalized_target = self._normalize(target_name)
            if normalized_target in already_mapped:
                raise ValidationError(
                    f"La columna {target_name!r} aparece más de una vez en el Excel."
                )
            already_mapped.add(normalized_target)
            indexes[index] = target_name

        if self.config.get("require_all_headers", True):
            required_names = self.field_names
        else:
            required_names = [
                str(field["name"])
                for field in self.fields
                if field.get("required", False)
            ]
        missing = [
            name
            for name in required_names
            if self._normalize(name) not in already_mapped
        ]
        if missing:
            raise ValidationError(
                "Faltan cabeceras requeridas en el Excel: " + ", ".join(missing)
            )
        if extra_headers and not self.config.get("ignore_extra_columns", True):
            raise ValidationError(
                "El Excel contiene cabeceras no configuradas: "
                + ", ".join(extra_headers)
            )
        return indexes, extra_headers

    def load(self, file_path: str) -> ExcelImportData:
        if load_workbook is None:
            raise ConfigurationError(
                "Falta openpyxl. Ejecuta: python -m pip install -r requirements.txt"
            )

        path = Path(file_path)
        if not path.is_file():
            raise ValidationError("El archivo Excel seleccionado no existe.")
        allowed = {
            str(extension).lower()
            for extension in self.config.get("allowed_extensions", [".xlsx", ".xlsm"])
        }
        if path.suffix.lower() not in allowed:
            raise ValidationError(
                "Formato no permitido. Usa: " + ", ".join(sorted(allowed))
            )

        workbook = load_workbook(
            filename=path,
            read_only=True,
            data_only=True,
            keep_links=False,
            keep_vba=path.suffix.lower() == ".xlsm",
        )
        try:
            worksheet = self._choose_sheet(workbook)
            row_iterator = worksheet.iter_rows(values_only=True)
            header_row_number = int(self.config.get("header_row", 1))
            for _ in range(header_row_number - 1):
                next(row_iterator, None)
            header_values = next(row_iterator, None)
            if header_values is None:
                raise ValidationError(
                    f"No se encontró la fila de cabeceras {header_row_number}."
                )
            indexes, _extra_headers = self._resolve_headers(header_values)

            preview_limit = int(self.config.get("preview_rows", 20))
            maximum = int(self.config.get("max_rows", 100000))
            preview_rows: list[dict[str, Any]] = []
            prepared_rows: list[list[Any]] = []
            source_rows: list[int] = []

            for excel_row_number, values in enumerate(
                row_iterator,
                start=header_row_number + 1,
            ):
                raw_row = {name: None for name in self.field_names}
                for index, target_name in indexes.items():
                    raw_row[target_name] = values[index] if index < len(values) else None
                if all(self._is_blank(value) for value in raw_row.values()):
                    continue
                if len(prepared_rows) >= maximum:
                    raise ValidationError(
                        f"El archivo supera el máximo configurado de {maximum} filas."
                    )

                try:
                    prepared = [
                        convert_value(raw_row[str(field["name"])], field)
                        for field in self.fields
                    ]
                except (ValidationError, ConfigurationError) as exc:
                    raise ValidationError(
                        f"Fila {excel_row_number} del Excel: {exc}"
                    ) from exc

                prepared_rows.append(prepared)
                source_rows.append(excel_row_number)
                if len(preview_rows) < preview_limit:
                    preview_rows.append(raw_row)

            if not prepared_rows:
                raise ValidationError("El archivo no contiene filas de datos para insertar.")

            return ExcelImportData(
                file_path=str(path),
                sheet_name=str(worksheet.title),
                columns=list(self.field_names),
                preview_rows=preview_rows,
                prepared_rows=prepared_rows,
                source_row_numbers=source_rows,
            )
        finally:
            workbook.close()
